# -*- coding: utf-8 -*-
"""Docgen service: LeAgent-aligned document/excel/slides/checklist/template generators."""
from __future__ import annotations

import json
import os
import re
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

from loguru import logger

from reactor_tool.docgen.paths import OUTPUT_DIR
from reactor_tool.util.file_util import upload_file_by_path, upload_file


def _safe_filename(name: str, default_ext: str) -> str:
    base = (name or "").strip() or f"docgen_{uuid.uuid4().hex[:8]}"
    base = os.path.basename(base.replace("\\", "/"))
    base = re.sub(r'[<>:"|?*]+', "_", base)
    if default_ext and not base.lower().endswith(default_ext.lower()):
        if "." not in Path(base).name:
            base = f"{base}{default_ext}"
    return base


def resolve_output_path(request_id: str, output_path: str | None, default_ext: str) -> Path:
    rid = (request_id or "default").strip() or "default"
    root = OUTPUT_DIR / rid
    root.mkdir(parents=True, exist_ok=True)
    if output_path:
        name = _safe_filename(output_path, default_ext)
    else:
        name = _safe_filename(f"output_{uuid.uuid4().hex[:8]}", default_ext)
    path = root / name
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


async def _upload_result(request_id: str, output_path: Path, result: dict[str, Any]) -> dict[str, Any]:
    file_info = None
    try:
        file_info = await upload_file_by_path(str(output_path), request_id=request_id)
    except Exception as exc:  # noqa: BLE001 - local-only envs may lack FILE_SERVER_URL
        logger.warning(f"docgen upload skipped path={output_path}: {exc}")
    payload = {
        "success": bool(result.get("success", True)),
        "outputPath": str(output_path),
        "fileInfo": [file_info] if file_info else [],
        "stats": result.get("stats") or result.get("content_stats") or {},
        "warnings": result.get("warnings") or [],
        "message": result.get("message") or "ok",
    }
    # keep original keys useful to agent
    for k in ("format", "sheet_names", "file_size_bytes", "rendered_length", "variables_used", "output_format"):
        if k in result:
            payload[k] = result[k]
    return payload


def generate_document(params: dict[str, Any], output_path: Path) -> dict[str, Any]:
    from reactor_tool.docgen.markdown import parse_markdown_document
    from reactor_tool.docgen.model import DocumentSpec, _coerce_date_str

    _FRONT_MATTER_KEYS = (
        "title", "subtitle", "author", "date", "subject", "keywords", "theme", "toc",
        "cover", "numbered_headings", "justify", "numbered_figures", "section_pages",
        "header", "footer", "watermark", "page",
    )
    _EXT_FORMATS = {".pdf": "pdf", ".docx": "docx", ".html": "html", ".htm": "html", ".md": "markdown", ".markdown": "markdown"}
    _FORMAT_EXTS = {"pdf": ".pdf", "docx": ".docx", "html": ".html", "markdown": ".md"}

    fmt = (params.get("format") or "").strip().lower()
    if not fmt:
        fmt = _EXT_FORMATS.get(output_path.suffix.lower(), "pdf")
    if fmt not in _FORMAT_EXTS:
        raise ValueError(f"Unsupported format: {fmt}")
    if output_path.suffix.lower() not in _EXT_FORMATS or _EXT_FORMATS.get(output_path.suffix.lower()) != fmt:
        output_path = output_path.with_suffix(_FORMAT_EXTS[fmt])

    blocks: list[Any] = []
    content = params.get("content")
    front_matter: dict[str, Any] = {}
    if isinstance(content, str) and content.strip():
        front_matter, parsed = parse_markdown_document(content)
        blocks.extend(parsed)
    raw_blocks = params.get("blocks")
    if isinstance(raw_blocks, list):
        blocks.extend(raw_blocks)
    if not blocks:
        raise ValueError("Provide non-empty `content` (markdown) and/or `blocks`.")

    work = dict(params)
    for key in _FRONT_MATTER_KEYS:
        if work.get(key) is None and key in front_matter:
            work[key] = front_matter[key]
    if work.get("date") is not None:
        work["date"] = _coerce_date_str(work["date"])

    spec = DocumentSpec.model_validate(
        {
            "title": work.get("title") or "",
            "subtitle": work.get("subtitle"),
            "author": work.get("author"),
            "date": work.get("date"),
            "subject": work.get("subject"),
            "keywords": work.get("keywords") or [],
            "theme": work.get("theme") or "professional",
            "toc": bool(work.get("toc")),
            "cover": work.get("cover", False),
            "numbered_headings": bool(work.get("numbered_headings")),
            "justify": bool(work.get("justify")),
            "numbered_figures": bool(work.get("numbered_figures")),
            "section_pages": bool(work.get("section_pages")),
            "header": work.get("header"),
            "footer": work.get("footer"),
            "watermark": work.get("watermark"),
            "page": work.get("page") or {},
            "encryption": work.get("encryption"),
            "merge_sources": work.get("merge_sources") or [],
            "blocks": blocks,
        }
    )
    logger.info(f"document_generate_start path={output_path} format={fmt} blocks={len(spec.blocks)}")
    if fmt == "pdf":
        from reactor_tool.docgen.renderers.pdf import render_pdf
        return render_pdf(spec, output_path)
    if fmt == "docx":
        from reactor_tool.docgen.renderers.docx import render_docx
        return render_docx(spec, output_path)
    if fmt == "html":
        from reactor_tool.docgen.renderers.html import render_html
        return render_html(spec, output_path)
    from reactor_tool.docgen.renderers.html import render_markdown
    return render_markdown(spec, output_path)


def generate_slides(params: dict[str, Any], output_path: Path) -> dict[str, Any]:
    from reactor_tool.docgen.model import DeckSpec
    from reactor_tool.docgen.renderers.pptx import render_pptx

    if output_path.suffix.lower() != ".pptx":
        output_path = output_path.with_suffix(".pptx")
    slides = params.get("slides")
    if not isinstance(slides, list) or not slides:
        raise ValueError("`slides` must be a non-empty array of slide objects.")
    deck = DeckSpec.model_validate(
        {
            "title": params.get("title") or "",
            "subtitle": params.get("subtitle"),
            "author": params.get("author"),
            "date": params.get("date"),
            "theme": params.get("theme") or "executive_light",
            "aspect": params.get("aspect") or "16:9",
            "show_slide_numbers": params.get("show_slide_numbers", True),
            "footer_text": params.get("footer_text"),
            "background": params.get("background"),
            "slides": slides,
        }
    )
    logger.info(f"slides_generate_start path={output_path} slides={len(deck.slides)}")
    return render_pptx(deck, output_path)


def generate_checklist(params: dict[str, Any], output_path: Path) -> dict[str, Any]:
    from reactor_tool.docgen.checklist import build_checklist_block, checklist_to_dict
    from reactor_tool.docgen.model import DocumentSpec

    _EXT_FORMATS = {
        ".pdf": "pdf", ".docx": "docx", ".html": "html", ".htm": "html",
        ".md": "markdown", ".markdown": "markdown", ".json": "json",
    }
    _FORMAT_EXTS = {"pdf": ".pdf", "docx": ".docx", "html": ".html", "markdown": ".md", "json": ".json"}
    fmt = (params.get("format") or params.get("output_format") or "").strip().lower()
    if not fmt:
        fmt = _EXT_FORMATS.get(output_path.suffix.lower(), "markdown")
    if fmt not in _FORMAT_EXTS:
        raise ValueError(f"Unsupported format: {fmt}")
    if _EXT_FORMATS.get(output_path.suffix.lower()) != fmt:
        output_path = output_path.with_suffix(_FORMAT_EXTS[fmt])

    block = build_checklist_block(params)
    if not block.normalized_groups() or not any(g.items for g in block.normalized_groups()):
        raise ValueError("Checklist is empty — provide `items`, `groups`, or a valid `source_path`.")

    if fmt == "json":
        payload = checklist_to_dict(block)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        return {"success": True, "output_path": str(output_path), "format": "json", "content_stats": payload["stats"], "warnings": []}

    spec = DocumentSpec.model_validate(
        {
            "title": block.title or params.get("title") or "Checklist",
            "theme": params.get("theme") or "professional",
            "blocks": [block],
        }
    )
    if fmt == "pdf":
        from reactor_tool.docgen.renderers.pdf import render_pdf
        return render_pdf(spec, output_path)
    if fmt == "docx":
        from reactor_tool.docgen.renderers.docx import render_docx
        return render_docx(spec, output_path)
    if fmt == "html":
        from reactor_tool.docgen.renderers.html import render_html
        return render_html(spec, output_path)
    from reactor_tool.docgen.renderers.html import render_markdown
    return render_markdown(spec, output_path)


def generate_excel(params: dict[str, Any], output_path: Path) -> dict[str, Any]:
    """Port of LeAgent ExcelGeneratorTool.execute_sync (without LibreOffice recalc)."""
    from openpyxl import Workbook
    from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart
    from openpyxl.chart.reference import Reference
    from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter, range_boundaries
    from openpyxl.worksheet.datavalidation import DataValidation

    if output_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        output_path = output_path.with_suffix(".xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    sheets_config = params.get("sheets") or []
    if not sheets_config:
        raise ValueError("`sheets` must be a non-empty array")
    wb_props = params.get("workbook_properties") or {}
    preset = params.get("preset")

    wb = Workbook()
    default_sheet = wb.active
    default_font_name = "Arial" if preset == "financial" else "Calibri"
    try:
        wb.style = default_font_name  # best-effort; openpyxl default font differs by version
    except Exception:
        pass

    if wb_props:
        if wb_props.get("title"):
            wb.properties.title = wb_props["title"]
        if wb_props.get("author"):
            wb.properties.creator = wb_props["author"]
        if wb_props.get("subject"):
            wb.properties.subject = wb_props["subject"]
        if wb_props.get("company"):
            wb.properties.company = wb_props["company"]

    stats = {"sheets": 0, "total_rows": 0, "total_cells": 0, "charts": 0, "formulas": 0, "merged_ranges": 0, "validations": 0}
    chart_classes = {"bar": BarChart, "column": BarChart, "line": LineChart, "pie": PieChart, "area": AreaChart}

    def apply_financial(ws, headers, data):
        blue_font = Font(color="0000FF", name="Arial", size=11)
        black_font = Font(color="000000", name="Arial", size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(bold=True, name="Arial", size=11)
        header_align = Alignment(horizontal="center")
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        start_row = 2 if headers else 1
        for row_idx, row_data in enumerate(data, start_row):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                is_formula = isinstance(value, str) and str(value).startswith("=")
                cell.font = black_font if is_formula else blue_font

    def create_dv(dv_def):
        dv_type = dv_def.get("type", "list")
        if dv_type == "list":
            formula1 = dv_def.get("formula1", "")
            if formula1 and not str(formula1).startswith('"'):
                formula1 = f'"{formula1}"'
            return DataValidation(type="list", formula1=formula1, allow_blank=True)
        if dv_type in ("whole", "decimal"):
            return DataValidation(
                type=dv_type,
                operator=dv_def.get("operator", "between"),
                formula1=dv_def.get("formula1"),
                formula2=dv_def.get("formula2"),
                allow_blank=True,
            )
        return None

    def apply_style(ws, style_def):
        range_str = style_def["range"]
        if ":" in range_str:
            min_col, min_row, max_col, max_row = range_boundaries(range_str)
        else:
            min_col, min_row, max_col, max_row = range_boundaries(f"{range_str}:{range_str}")
        font_kwargs: dict[str, Any] = {}
        if style_def.get("bold"):
            font_kwargs["bold"] = True
        if style_def.get("italic"):
            font_kwargs["italic"] = True
        if style_def.get("font_size"):
            font_kwargs["size"] = style_def["font_size"]
        if style_def.get("font_name"):
            font_kwargs["name"] = style_def["font_name"]
        if style_def.get("font_color"):
            font_kwargs["color"] = style_def["font_color"]
        fill = None
        if style_def.get("bg_color"):
            fill = PatternFill(start_color=style_def["bg_color"], end_color=style_def["bg_color"], fill_type="solid")
        align = None
        if style_def.get("align") or style_def.get("vertical") or style_def.get("wrap_text"):
            align = Alignment(
                horizontal=style_def.get("align"),
                vertical=style_def.get("vertical"),
                wrap_text=bool(style_def.get("wrap_text")),
            )
        border = None
        if style_def.get("border"):
            side = Side(style=style_def.get("border_style") or "thin", color=style_def.get("border_color") or "000000")
            border = Border(left=side, right=side, top=side, bottom=side)
        font = Font(**font_kwargs) if font_kwargs else None
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill
                if align:
                    cell.alignment = align
                if border:
                    cell.border = border
                if style_def.get("number_format"):
                    cell.number_format = style_def["number_format"]

    def apply_cf(ws, cf_def):
        rng = cf_def["range"]
        rule_type = cf_def.get("type") or cf_def.get("rule_type") or "cell_value"
        fill_color = cf_def.get("fill_color") or cf_def.get("color") or "FFFF00"
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        if rule_type in ("color_scale", "colour_scale"):
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ))
            return
        if rule_type == "data_bar":
            ws.conditional_formatting.add(rng, DataBarRule(start_type="min", end_type="max", color=fill_color))
            return
        op = cf_def.get("operator") or "greaterThan"
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator=op, formula=[str(cf_def.get("value", "0"))], fill=fill),
        )

    for idx, sheet_config in enumerate(sheets_config):
        sheet_name = str(sheet_config.get("name") or f"Sheet{idx + 1}")[:31]
        if idx == 0 and default_sheet is not None:
            ws = default_sheet
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)
        headers = sheet_config.get("headers") or []
        data = sheet_config.get("data") or []
        if headers:
            header_font = Font(bold=True, name=default_font_name)
            header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
            header_align = Alignment(horizontal="center")
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                stats["total_cells"] += 1
            stats["total_rows"] += 1
        start_row = 2 if headers else 1
        for row_idx, row_data in enumerate(data, start_row):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
                stats["total_cells"] += 1
            stats["total_rows"] += 1
        if preset == "financial" and (headers or data):
            apply_financial(ws, headers, data)
        for col_letter, width in (sheet_config.get("column_widths") or {}).items():
            ws.column_dimensions[str(col_letter)].width = width
        for row_str, height in (sheet_config.get("row_heights") or {}).items():
            try:
                ws.row_dimensions[int(row_str)].height = height
            except (ValueError, TypeError):
                pass
        if sheet_config.get("freeze_panes"):
            ws.freeze_panes = sheet_config["freeze_panes"]
        if sheet_config.get("auto_filter") and (headers or data):
            max_row = len(data) + (1 if headers else 0)
            max_col = len(headers) if headers else (len(data[0]) if data else 0)
            if max_col > 0:
                ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
        for formula_def in sheet_config.get("formulas") or []:
            ws[formula_def["cell"]] = formula_def["formula"]
            stats["formulas"] += 1
        for merge_range in sheet_config.get("merged_cells") or []:
            ws.merge_cells(merge_range)
            stats["merged_ranges"] += 1
        for dv_def in sheet_config.get("data_validation") or []:
            dv = create_dv(dv_def)
            if dv:
                ws.add_data_validation(dv)
                dv.add(dv_def["range"])
                stats["validations"] += 1
        for style_def in sheet_config.get("cell_styles") or []:
            apply_style(ws, style_def)
        for chart_def in sheet_config.get("charts") or []:
            chart_type = chart_def.get("type")
            chart_class = chart_classes.get(chart_type)
            if not chart_class:
                continue
            chart = chart_class()
            if chart_def.get("title"):
                chart.title = chart_def["title"]
            min_col, min_row, max_col, max_row = range_boundaries(chart_def["data_range"])
            data_ref = Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
            chart.add_data(data_ref, titles_from_data=True)
            if chart_def.get("categories_range"):
                cmin_col, cmin_row, cmax_col, cmax_row = range_boundaries(chart_def["categories_range"])
                chart.set_categories(Reference(ws, min_col=cmin_col, min_row=cmin_row, max_col=cmax_col, max_row=cmax_row))
            if chart_type == "column":
                chart.type = "col"
            if chart_def.get("width"):
                chart.width = chart_def["width"]
            if chart_def.get("height"):
                chart.height = chart_def["height"]
            ws.add_chart(chart, chart_def.get("position") or "E2")
            stats["charts"] += 1
        for cf_def in sheet_config.get("conditional_formatting") or []:
            apply_cf(ws, cf_def)
        stats["sheets"] += 1

    wb.save(str(output_path))
    return {
        "success": True,
        "output_path": str(output_path),
        "file_size_bytes": output_path.stat().st_size,
        "stats": stats,
        "sheet_names": [str(s.get("name") or f"Sheet{i+1}")[:31] for i, s in enumerate(sheets_config)],
    }


def fill_template(params: dict[str, Any], output_path: Path | None = None) -> dict[str, Any]:
    from jinja2 import (
        BaseLoader, ChoiceLoader, Environment, FileSystemLoader,
        StrictUndefined, Undefined, select_autoescape,
    )
    from jinja2.exceptions import TemplateError, TemplateSyntaxError, UndefinedError

    template_source = params.get("template_source") or "string"
    strict_mode = bool(params.get("strict_mode", False))
    autoescape = params.get("autoescape", True)
    trim_blocks = params.get("trim_blocks", True)
    lstrip_blocks = params.get("lstrip_blocks", True)
    output_format = params.get("output_format") or "text"

    variables: dict[str, Any] = {}
    if params.get("data_file"):
        data_path = Path(params["data_file"])
        content = data_path.read_text(encoding="utf-8")
        if data_path.suffix in (".yaml", ".yml"):
            import yaml
            file_vars = yaml.safe_load(content)
        elif data_path.suffix == ".json":
            file_vars = json.loads(content)
        else:
            raise ValueError(f"Unsupported data file format: {data_path.suffix}")
        if isinstance(file_vars, dict):
            variables.update(file_vars)
    if params.get("variables"):
        variables.update(params["variables"])

    loaders: list[BaseLoader] = []
    if template_source == "file" and params.get("template_path"):
        template_path = Path(params["template_path"])
        if not template_path.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")
        loaders.append(FileSystemLoader(str(template_path.parent)))
    for include_path in params.get("include_paths") or []:
        path = Path(include_path)
        if path.exists() and path.is_dir():
            loaders.append(FileSystemLoader(str(path)))
    loader = ChoiceLoader(loaders) if loaders else None
    undefined_class = StrictUndefined if strict_mode else Undefined
    if autoescape:
        autoescape_config = select_autoescape(
            enabled_extensions=("html", "htm", "xml"),
            default_for_string=output_format in ("html", "xml"),
        )
    else:
        autoescape_config = False
    env = Environment(
        loader=loader,
        undefined=undefined_class,
        autoescape=autoescape_config,
        trim_blocks=trim_blocks,
        lstrip_blocks=lstrip_blocks,
    )

    # builtin filters
    def format_date(value: Any, fmt: str = "%Y-%m-%d") -> str:
        if isinstance(value, str):
            try:
                value = datetime.fromisoformat(value)
            except ValueError:
                return str(value)
        if isinstance(value, datetime):
            return value.strftime(fmt)
        return str(value)

    def format_currency(value: Any, symbol: str = "$", decimals: int = 2) -> str:
        try:
            return f"{symbol}{float(value):,.{decimals}f}"
        except (ValueError, TypeError):
            return str(value)

    def slugify(value: str) -> str:
        value = value.lower().strip()
        value = re.sub(r"[^\w\s-]", "", value)
        return re.sub(r"[-\s]+", "-", value)

    env.filters["format_date"] = format_date
    env.filters["format_currency"] = format_currency
    env.filters["slugify"] = slugify
    env.globals["now"] = datetime.now
    for name, value in (params.get("globals") or {}).items():
        env.globals[name] = value

    if template_source == "file":
        template = env.get_template(Path(params["template_path"]).name)
    elif template_source == "string":
        template_string = params.get("template_string") or ""
        if not template_string:
            raise ValueError("template_string is required for 'string' source")
        template = env.from_string(template_string)
    elif template_source == "url":
        import requests
        template_url = params.get("template_url")
        if not template_url:
            raise ValueError("template_url is required for 'url' source")
        resp = requests.get(template_url, timeout=30)
        resp.raise_for_status()
        template = env.from_string(resp.text)
    else:
        raise ValueError(f"Unknown template source: {template_source}")

    try:
        rendered = template.render(**variables)
    except UndefinedError as e:
        raise ValueError(f"Undefined variable in template: {e}") from e
    except TemplateSyntaxError as e:
        raise ValueError(f"Template syntax error at line {e.lineno}: {e.message}") from e
    except TemplateError as e:
        raise RuntimeError(f"Template rendering error: {e}") from e

    result: dict[str, Any] = {
        "success": True,
        "rendered": rendered,
        "rendered_length": len(rendered),
        "variables_used": list(variables.keys()),
        "output_format": output_format,
    }
    if output_path is not None:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(rendered, encoding="utf-8")
        result["output_path"] = str(output_path)
        result["file_size_bytes"] = output_path.stat().st_size
    return result


async def run_document_generate(request_id: str, params: dict[str, Any]) -> dict[str, Any]:
    ext_map = {"pdf": ".pdf", "docx": ".docx", "html": ".html", "markdown": ".md"}
    fmt = (params.get("format") or "").strip().lower() or "pdf"
    path = resolve_output_path(request_id, params.get("output_path") or params.get("fileName"), ext_map.get(fmt, ".pdf"))
    result = generate_document(params, path)
    # renderers may write a different final path
    final = Path(result.get("output_path") or path)
    return await _upload_result(request_id, final, result)


async def run_slides_generate(request_id: str, params: dict[str, Any]) -> dict[str, Any]:
    path = resolve_output_path(request_id, params.get("output_path") or params.get("fileName"), ".pptx")
    result = generate_slides(params, path)
    final = Path(result.get("output_path") or path)
    return await _upload_result(request_id, final, result)


async def run_excel_generator(request_id: str, params: dict[str, Any]) -> dict[str, Any]:
    path = resolve_output_path(request_id, params.get("output_path") or params.get("fileName"), ".xlsx")
    result = generate_excel(params, path)
    final = Path(result.get("output_path") or path)
    return await _upload_result(request_id, final, result)


async def run_checklist_generate(request_id: str, params: dict[str, Any]) -> dict[str, Any]:
    fmt = (params.get("format") or params.get("output_format") or "markdown").strip().lower()
    ext = {".pdf": "pdf", "pdf": ".pdf", "docx": ".docx", "html": ".html", "markdown": ".md", "json": ".json"}.get(fmt, ".md")
    if not ext.startswith("."):
        ext = f".{ext}"
    path = resolve_output_path(request_id, params.get("output_path") or params.get("fileName"), ext)
    result = generate_checklist(params, path)
    final = Path(result.get("output_path") or path)
    return await _upload_result(request_id, final, result)


async def run_template_filler(request_id: str, params: dict[str, Any]) -> dict[str, Any]:
    out_name = params.get("output_path") or params.get("fileName")
    path = resolve_output_path(request_id, out_name, ".txt") if out_name else resolve_output_path(request_id, "template_output.txt", ".txt")
    result = fill_template(params, path)
    final = Path(result.get("output_path") or path)
    # also return rendered text for agent
    payload = await _upload_result(request_id, final, result)
    payload["rendered"] = result.get("rendered", "")
    return payload


def _deep_merge(base: dict[str, Any], override: dict[str, Any]) -> None:
    for key, value in override.items():
        if key in base and isinstance(base[key], dict) and isinstance(value, dict):
            _deep_merge(base[key], value)
        else:
            base[key] = value


async def run_document_template(request_id: str, params: dict[str, Any]) -> dict[str, Any]:
    """Manage reusable docgen templates (save/list/get/delete/preview/generate)."""
    from reactor_tool.docgen import templates as store

    action = str(params.get("action") or "").strip().lower()
    actions = ("save", "list", "get", "delete", "preview", "generate")
    if action not in actions:
        raise ValueError(f"Unknown action: {action!r}. Use one of {actions}.")

    if action == "list":
        return {"success": True, "message": "ok", "templates": store.list_templates(), "fileInfo": []}

    name = params.get("name")
    if not name:
        raise ValueError(f"`name` is required for action={action}.")

    if action == "save":
        template = store.DocTemplate.model_validate(
            {
                "name": name,
                "kind": params.get("kind") or "document",
                "description": params.get("description"),
                "theme": params.get("theme"),
                "variables": params.get("variables") or [],
                "content": params.get("content"),
                "slides": params.get("slides"),
                "defaults": params.get("defaults") or {},
            }
        )
        saved = store.save_template(template, overwrite=params.get("overwrite", True))
        saved["success"] = True
        saved["message"] = "ok"
        saved["fileInfo"] = []
        saved["usage"] = (
            f"Instantiate with action='generate', name='{saved['name']}', "
            "values={{...}}, output_path='...'."
        )
        return saved

    if action == "delete":
        removed = store.delete_template(str(name))
        return {"success": True, "message": "ok", "deleted": removed, "name": name, "fileInfo": []}

    template = store.load_template(str(name))
    if template is None:
        raise ValueError(f"Template not found: {name!r}")

    if action == "get":
        return {
            "success": True,
            "message": "ok",
            "template": template.model_dump(exclude_none=True),
            "fileInfo": [],
        }

    payload = store.render_template(template, params.get("values") or {})
    if action == "preview":
        return {"success": True, "message": "ok", "name": template.name, "rendered": payload, "fileInfo": []}

    # generate
    output_path = params.get("output_path") or params.get("fileName")
    if not output_path:
        raise ValueError("`output_path` is required for action=generate.")
    payload.pop("kind", None)
    if template.kind == "deck":
        path = resolve_output_path(request_id, output_path, ".pptx")
        result = generate_slides({**payload, "output_path": str(path)}, path)
        final = Path(result.get("output_path") or path)
        return await _upload_result(request_id, final, result)

    fmt = (params.get("format") or payload.get("format") or "pdf").strip().lower()
    ext_map = {"pdf": ".pdf", "docx": ".docx", "html": ".html", "markdown": ".md"}
    path = resolve_output_path(request_id, output_path, ext_map.get(fmt, ".pdf"))
    gen_params = {**payload, "output_path": str(path), "format": fmt}
    result = generate_document(gen_params, path)
    final = Path(result.get("output_path") or path)
    return await _upload_result(request_id, final, result)


async def run_theme_designer(request_id: str, params: dict[str, Any]) -> dict[str, Any]:
    """Create/list/get/delete custom docgen themes."""
    del request_id  # no file upload for theme metadata ops
    from reactor_tool.docgen import theming
    from reactor_tool.docgen.themes import BUILTIN_THEMES, get_theme

    action = str(params.get("action") or "").strip().lower()
    actions = ("create", "save", "list", "get", "delete")
    if action not in actions:
        raise ValueError(f"Unknown action: {action!r}. Use one of {actions}.")
    kind = params.get("kind") or "document"
    name = params.get("name")

    if action == "list":
        builtin = [{"name": t.name, "kind": t.kind, "builtin": True} for t in BUILTIN_THEMES.values()]
        custom = [{**item, "builtin": False} for item in theming.list_custom_themes()]
        return {"success": True, "message": "ok", "themes": builtin + custom, "fileInfo": []}

    if not name:
        raise ValueError(f"`name` is required for action={action}.")

    if action == "get":
        resolved = get_theme(str(name), kind=kind)
        payload = theming.load_custom_theme_payload(str(name))
        return {
            "success": True,
            "message": "ok",
            "name": resolved.name,
            "builtin": str(name) in BUILTIN_THEMES,
            "payload": payload,
            "resolved": resolved.model_dump(),
            "lint_warnings": theming.lint_theme(resolved),
            "fileInfo": [],
        }

    if action == "delete":
        removed = theming.delete_custom_theme(str(name))
        return {"success": True, "message": "ok", "deleted": removed, "name": name, "fileInfo": []}

    if action == "create":
        primary = params.get("primary")
        if not primary:
            raise ValueError("`primary` (brand color '#RRGGBB') is required for create.")
        payload = theming.derive_theme_payload(
            kind=kind,
            primary=str(primary),
            accent=params.get("accent"),
            mode=params.get("mode"),
            heading_font=params.get("heading_font"),
            body_font=params.get("body_font"),
            east_asia_font=params.get("east_asia_font"),
        )
        overrides = params.get("overrides")
        if isinstance(overrides, dict) and overrides:
            _deep_merge(payload, overrides)
    else:  # save
        payload = params.get("payload")
        if not isinstance(payload, dict) or not payload:
            raise ValueError("`payload` (theme object) is required for save.")

    resolved = get_theme({**payload, "name": "candidate"}, kind=kind)
    if resolved.name != "candidate":
        raise ValueError("Theme payload failed validation against the theme schema.")
    lint = theming.lint_theme(resolved)
    result: dict[str, Any] = {
        "success": True,
        "message": "ok",
        "name": name,
        "kind": kind,
        "payload": payload,
        "colors": resolved.colors.model_dump(),
        "lint_warnings": lint,
        "fileInfo": [],
    }
    if params.get("dry_run"):
        result["saved"] = False
        return result

    saved = theming.save_custom_theme(str(name), payload, kind=kind)
    result.update({"saved": True, "path": saved["path"]})
    result["usage"] = (
        f'Pass theme: "{saved["name"]}" to '
        + ("slides_generate" if kind == "deck" else "document_generate")
        + "."
    )
    return result


async def run_chart_generator(request_id: str, params: dict[str, Any]) -> dict[str, Any]:
    from reactor_tool.tool.docgen.chart_generator import generate_chart

    fmt = (params.get("output_format") or "png").strip().lower()
    path = resolve_output_path(request_id, params.get("output_path") or params.get("fileName"), f".{fmt}")
    result = generate_chart(params, path)
    final = Path(result.get("output_path") or path)
    payload = await _upload_result(request_id, final, result)
    for k in ("chart_type", "theme", "format"):
        if k in result:
            payload[k] = result[k]
    return payload
